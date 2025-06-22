from flask import Flask, request, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ExifTags
from openpyxl.styles import Font, Alignment
import tempfile
import os
import json
from datetime import datetime, timedelta

app = Flask(__name__)

@app.route('/generar_excel', methods=['POST'])
def generar_excel():
    try:
        datos = request.form.get('datos')  
        rubros = json.loads(datos)
        files = request.files

        wb = Workbook()
        ws = wb.active

        temp_imgs = []

        nombre_instalacion = request.form.get('instalacion', 'Instalación')
        #  hora  zona 
        hora_venezuela = datetime.utcnow() - timedelta(hours=4)
        fecha_hora = hora_venezuela.strftime('%d/%m/%Y %H:%M')
        titulo = f"{nombre_instalacion} - {fecha_hora}"
         
        usuario = request.form.get('usuario', 'Usuario desconocido')

        # Usuario
        ws.merge_cells('A1:H1')
        cell_usuario = ws['A1']
        cell_usuario.value = f" {usuario}"
        cell_usuario.font = Font(size=14, italic=True)
        cell_usuario.alignment = Alignment(horizontal='center', vertical='center')

        # Título 
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = titulo
        cell.font = Font(size=18, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Encabezados 
        ws.append(['Nivel', 'Área', 'Rubro', 'Unidad\nResponsable', 'Desviación', 'Criticidad', 'Estatus', 'Foto'])

        for i, rubro in enumerate(rubros):
            fila = [
                rubro['nivel'],
                rubro['area'],
                rubro['rubro'],
                rubro['unidad_responsable'],
                rubro['desviacion'],
                rubro.get('criticidad', ''),
                'Solventado' if rubro['estatus'] else 'No solventado'
            ]
            ws.append(fila)
            ws.row_dimensions[i+4].height = 120

            if f"imagen_{i}" in files:
                img_file = files[f"imagen_{i}"]
                temp_img = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                img_file.save(temp_img.name)
                temp_imgs.append(temp_img.name)

              
                with PILImage.open(temp_img.name) as im:
                    try:
                        for orientation in ExifTags.TAGS.keys():
                            if ExifTags.TAGS[orientation] == 'Orientation':
                                break
                        exif = im._getexif()
                        if exif is not None:
                            orientation_value = exif.get(orientation, None)
                            if orientation_value == 3:
                                im = im.rotate(180, expand=True)
                            elif orientation_value == 6:
                                im = im.rotate(270, expand=True)
                            elif orientation_value == 8:
                                im = im.rotate(90, expand=True)
                    except Exception:
                        pass
                    im.thumbnail((500, 500))
                    im.save(temp_img.name, format='PNG')

                img = XLImage(temp_img.name)
                img.width = 110
                img.height = 160
                ws.add_image(img, f'H{i+4}')  # Cambia la columna a H (8va columna)

        # Encabezados 
        for col in range(1, 9):  # Ahora son 8 columnas
            c = ws.cell(row=3, column=col)
            c.font = Font(size=14, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')

        #  columnas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8  
        ws.column_dimensions['C'].width = 10  
        ws.column_dimensions['D'].width = 12  
        ws.column_dimensions['E'].width = 14  
        ws.column_dimensions['F'].width = 10  
        ws.column_dimensions['G'].width = 12  
        ws.column_dimensions['H'].width = 15  

       
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
            for idx, cell in enumerate(row):
                cell.alignment = Alignment(
                    vertical='center',
                    horizontal='center',
                    wrap_text=True if idx in [0, 1, 2, 3, 4, 5, 6] else False  # Rubro, Unidad Responsable, Área, Desviación, Criticidad, Estatus
                )

        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_excel.name)
        temp_excel.close()


        for temp_img_path in temp_imgs:
            try:
                os.unlink(temp_img_path)
            except Exception:
                pass

        return send_file(temp_excel.name, as_attachment=True, download_name=titulo + '.xlsx')
    except Exception as e:
        print("ERROR AL GENERAR EXCEL:", e)
        return "Error interno del servidor: " + str(e), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)